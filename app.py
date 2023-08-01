import openpyxl

def create_excel_sheet(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Student Detials"
    sheet["A1"] = "Student Name"
    sheet["B1"] = "Student Age"
    sheet["C1"] = "Gender"
    sheet["D1"] = "present days"
    sheet["E1"] = "total days"
    workbook.save(file_name)

def append_to_excel(file_name, name, age, gender, present_days, total_days):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook["Student Detials"]  
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=name)
    sheet.cell(row=next_row, column=2, value=age)
    sheet.cell(row=next_row, column=3, value=gender)
    sheet.cell(row=next_row, column=4, value=present_days)
    sheet.cell(row=next_row, column=5, value=total_days)
    workbook.save(file_name)


def main():
    file_name = "student.xlsx"
    try:
        create_excel_sheet(file_name)
    except Exception as e:
        print("Error creating Excel sheet:", e)
        return

    while True:
        name = input("Enter student Name: (or 'exit' to stop): ")
        if name.lower() == "exit":
            break

        age =int(input("Enter age: "))
        gender=input("Enter Gender:")
        present_days=input("Enter Present_days:")
        total_days=input("Enter Total_days:")

        try:
            append_to_excel(file_name, name, age, gender, present_days, total_days)
            print("Successfully added data to the Excel sheet.")
        except Exception as e:
            print("Error appending data to Excel sheet:", e)

if __name__ == "__main__":
    main()